"""
XLSX → WorkbookSnapshot extraction.

Walks an XLSX file with openpyxl (non-read-only mode, because we need
merged cell ranges and defined names) and produces a fully populated
WorkbookSnapshot — minus blocks, which are filled in by block_detector
in the next stage.

Typical usage:

    from excel_evo_tracker.extractor import extract_workbook
    snapshot = extract_workbook(Path("cache/xlsx/month_01.xlsx"))
"""

from __future__ import annotations

import datetime as dt
import logging
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from . import config
from .converter import compute_file_hash
from .models import (
    Cell,
    CellType,
    NamedRange,
    SheetSnapshot,
    WorkbookSnapshot,
)

logger = logging.getLogger(__name__)


# ── Cell classification ───────────────────────────────────────────────


def _classify_value(value) -> CellType:
    """Infer a CellType from a raw openpyxl cell value."""
    if value is None or value == "":
        return CellType.EMPTY
    if isinstance(value, bool):
        return CellType.BOOLEAN
    if isinstance(value, (int, float)):
        return CellType.NUMERIC
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return CellType.DATE
    # Everything else is treated as text (strings, or stringified oddities)
    return CellType.LABEL


def _build_merge_lookup(ws: Worksheet) -> dict[str, str]:
    """
    Map each cell address inside a merged region to the range string.

    Example: for a merge A1:C1, returns
        {"A1": "A1:C1", "B1": "A1:C1", "C1": "A1:C1"}
    """
    lookup: dict[str, str] = {}
    for merged in ws.merged_cells.ranges:
        range_str = str(merged)
        for row in ws[range_str]:
            for cell in row:
                lookup[cell.coordinate] = range_str
    return lookup


# ── Sheet extraction ──────────────────────────────────────────────────


def _extract_sheet(ws: Worksheet, index: int) -> SheetSnapshot:
    """Extract a single worksheet into a SheetSnapshot (without blocks)."""
    sheet_name = ws.title

    # Merged regions
    merged_regions = [str(r) for r in ws.merged_cells.ranges]
    merge_lookup = _build_merge_lookup(ws)

    # Determine used range
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    used_range = None
    if max_row > 0 and max_col > 0:
        used_range = f"A1:{get_column_letter(max_col)}{max_row}"

    # Warn on very large sheets
    total_possible = max_row * max_col
    if total_possible > config.LARGE_SHEET_WARN_THRESHOLD:
        logger.warning(
            "Sheet %r is large (%d × %d = %d cells); extraction may be slow.",
            sheet_name, max_row, max_col, total_possible,
        )

    # Walk cells, build sparse map
    cells: dict[str, Cell] = {}
    if max_row > 0 and max_col > 0:
        for row in ws.iter_rows(
            min_row=1, max_row=max_row,
            min_col=1, max_col=max_col,
        ):
            for xl_cell in row:
                dtype = _classify_value(xl_cell.value)
                if dtype == CellType.EMPTY:
                    continue
                if dtype == CellType.NUMERIC and not config.INCLUDE_NUMERIC_CELLS:
                    continue

                address = xl_cell.coordinate
                merge_range = merge_lookup.get(address)

                # Normalize value: labels stored as stripped strings
                value = xl_cell.value
                if dtype == CellType.LABEL:
                    value = str(value).strip()
                    if not value:  # became empty after stripping
                        continue
                elif dtype == CellType.DATE:
                    value = value.isoformat() if hasattr(value, "isoformat") else str(value)

                cells[address] = Cell(
                    address=address,
                    row=xl_cell.row,
                    col=xl_cell.column,
                    value=value,
                    dtype=dtype,
                    is_merged=merge_range is not None,
                    merge_range=merge_range,
                )

    return SheetSnapshot(
        name=sheet_name,
        index=index,
        is_hidden=ws.sheet_state != "visible",
        used_range=used_range,
        max_row=max_row,
        max_col=max_col,
        cells=cells,
        blocks=[],            # filled by block_detector
        merged_regions=merged_regions,
    )


# ── Named ranges ──────────────────────────────────────────────────────


def _extract_named_ranges(wb) -> list[NamedRange]:
    """Extract all defined names from the workbook."""
    result: list[NamedRange] = []

    # Workbook-scoped names
    for name_obj in wb.defined_names.values():
        try:
            refers_to = name_obj.value or ""
        except Exception as e:
            logger.warning("Could not read defined_name %r: %s", name_obj.name, e)
            refers_to = ""
        result.append(NamedRange(
            name=name_obj.name,
            scope="workbook",
            refers_to=refers_to,
        ))

    # Sheet-scoped names
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # openpyxl exposes sheet-scoped names via ws.defined_names
        scoped = getattr(ws, "defined_names", None)
        if scoped is None:
            continue
        try:
            iterable = scoped.values() if hasattr(scoped, "values") else scoped
        except Exception:
            continue
        for name_obj in iterable:
            try:
                refers_to = name_obj.value or ""
            except Exception:
                refers_to = ""
            result.append(NamedRange(
                name=name_obj.name,
                scope=sheet_name,
                refers_to=refers_to,
            ))

    return result


# ── Public API ────────────────────────────────────────────────────────


def extract_workbook(
    xlsx_path: Path,
    *,
    original_xlsb_path: Path | None = None,
    month_label: str | None = None,
) -> WorkbookSnapshot:
    """
    Extract a complete structural snapshot of an XLSX workbook.

    Args:
        xlsx_path: Path to the XLSX file (typically a converted XLSB).
        original_xlsb_path: Optional path to the source XLSB, used for
            the file_hash field (so downstream dedup is keyed on the
            original binary, not the converted XLSX).
        month_label: Optional user-assigned label, e.g. "2024-01".

    Returns:
        A WorkbookSnapshot with all sheets, cells, merged regions, and
        named ranges populated. Blocks are empty — they're added by
        block_detector.detect_blocks_for_workbook().
    """
    xlsx_path = Path(xlsx_path).resolve()
    if not xlsx_path.exists():
        raise FileNotFoundError(f"XLSX not found: {xlsx_path}")

    # Hash the original XLSB if provided, else hash the XLSX
    hash_source = Path(original_xlsb_path) if original_xlsb_path else xlsx_path
    file_hash = compute_file_hash(hash_source)

    logger.info("Extracting %s", xlsx_path.name)
    wb = load_workbook(
        filename=str(xlsx_path),
        read_only=False,     # need merges, defined names
        data_only=True,      # cached values, not formula text
        keep_links=False,
    )

    try:
        sheets: dict[str, SheetSnapshot] = {}
        sheet_names: list[str] = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]

            if ws.sheet_state != "visible" and not config.INCLUDE_HIDDEN_SHEETS:
                logger.debug("Skipping hidden sheet %r", sheet_name)
                continue

            sheet_snap = _extract_sheet(ws, index=idx)
            sheets[sheet_name] = sheet_snap
            sheet_names.append(sheet_name)

        named_ranges = _extract_named_ranges(wb)

        total_cells = sum(len(s.cells) for s in sheets.values())

        snapshot = WorkbookSnapshot(
            file_name=(original_xlsb_path or xlsx_path).name,
            file_path=str((original_xlsb_path or xlsx_path).resolve()),
            file_hash=file_hash,
            snapshot_date=dt.datetime.now(dt.timezone.utc).isoformat(),
            month_label=month_label,
            sheet_names=sheet_names,
            sheets=sheets,
            named_ranges=named_ranges,
            total_blocks=0,   # updated after block detection
            total_cells=total_cells,
            metadata={
                "xlsx_path": str(xlsx_path),
                "sheet_count": len(sheet_names),
                "named_range_count": len(named_ranges),
            },
        )

        logger.info(
            "Extracted %s: %d sheets, %d cells, %d named ranges",
            xlsx_path.name, len(sheet_names), total_cells, len(named_ranges),
        )
        return snapshot

    finally:
        wb.close()
