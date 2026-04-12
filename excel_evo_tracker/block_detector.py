"""
Block detection via connected component analysis.

Turns a sparse cell map into a list of semantic blocks — spatially
contiguous clusters of cells that form logical units (label-value
groups, small tables, header regions).

Algorithm:
  1. Build a binary occupancy grid from the sheet's non-empty cells.
  2. Run 8-connectivity labeling with scipy.ndimage.label to get tight
     components (no gap tolerance at this stage).
  3. Post-hoc: merge component pairs whose bounding boxes are separated
     by at most GAP_TOLERANCE empty rows/cols AND overlap in the
     perpendicular dimension. This gives precise control — a tolerance
     of N means literally "up to N empty rows or cols between blocks".
  4. For each merged component, collect the original cells inside its
     bounding box, and emit a Block with labels, primary_label,
     fingerprint, and neighbor_context.

A debug visualizer renders the detected blocks as ASCII grids to a file
so you can eyeball clustering quality on real sheets before committing
to matcher/differ tuning.

Typical usage:

    from excel_evo_tracker.block_detector import detect_blocks_for_workbook
    detect_blocks_for_workbook(snapshot)           # mutates in place
"""

from __future__ import annotations

import hashlib
import logging
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from scipy.ndimage import label as nd_label

from . import config
from .models import Block, Cell, CellType, SheetSnapshot, WorkbookSnapshot

logger = logging.getLogger(__name__)


# ── Internal helpers ──────────────────────────────────────────────────


@dataclass
class _ComponentBounds:
    """Bounding box of a single connected component (0-indexed, inclusive)."""
    top: int
    left: int
    bottom: int
    right: int

    @property
    def row_span(self) -> int:
        return self.bottom - self.top + 1

    @property
    def col_span(self) -> int:
        return self.right - self.left + 1


def _build_occupancy_grid(sheet: SheetSnapshot) -> np.ndarray:
    """
    Build a binary grid: 1 where a cell is non-empty, 0 elsewhere.

    Returned shape is (max_row, max_col) with 0-indexed internal
    coordinates. Excel addresses are 1-indexed, so we subtract 1 when
    placing cells into the grid.

    Merged regions are filled in across their full range. Openpyxl only
    stores the value at the top-left of a merge, but semantically the
    entire merged range is "occupied" — treating it that way keeps
    block detection from artificially splitting around merged titles
    and headers.
    """
    if sheet.max_row == 0 or sheet.max_col == 0:
        return np.zeros((0, 0), dtype=np.uint8)

    grid = np.zeros((sheet.max_row, sheet.max_col), dtype=np.uint8)
    for cell in sheet.cells.values():
        r = cell.row - 1
        c = cell.col - 1
        if 0 <= r < sheet.max_row and 0 <= c < sheet.max_col:
            grid[r, c] = 1

    # Fill merged regions across their full span
    from openpyxl.utils.cell import range_boundaries
    for merge_range in sheet.merged_regions:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(merge_range)
        except Exception as e:
            logger.warning("Could not parse merge range %r: %s", merge_range, e)
            continue
        r_start = max(0, min_row - 1)
        r_end = min(sheet.max_row, max_row)
        c_start = max(0, min_col - 1)
        c_end = min(sheet.max_col, max_col)
        if r_start < r_end and c_start < c_end:
            grid[r_start:r_end, c_start:c_end] = 1

    return grid


def _compute_component_bounds(
    labeled: np.ndarray, num_components: int
) -> dict[int, _ComponentBounds]:
    """Compute bounding boxes for every labeled component."""
    bounds: dict[int, _ComponentBounds] = {}
    if num_components == 0:
        return bounds

    # Vectorized bounding box computation
    for comp_id in range(1, num_components + 1):
        rows, cols = np.where(labeled == comp_id)
        if rows.size == 0:
            continue
        bounds[comp_id] = _ComponentBounds(
            top=int(rows.min()),
            left=int(cols.min()),
            bottom=int(rows.max()),
            right=int(cols.max()),
        )
    return bounds


def _pick_primary_label(label_cells: list[Cell]) -> str | None:
    """
    Choose the best identity label for a block.

    Heuristic: top-most row, then left-most column. Ties broken by the
    longer string (more descriptive labels win). If the block has no
    label cells, returns None — caller should skip or demote the block.
    """
    if not label_cells:
        return None
    sorted_cells = sorted(
        label_cells,
        key=lambda c: (c.row, c.col, -len(str(c.value or ""))),
    )
    return str(sorted_cells[0].value) if sorted_cells[0].value is not None else None


def _compute_fingerprint(
    label_cells: list[Cell],
    top_row: int,
    top_col: int,
) -> str:
    """
    Position-independent structural hash of a block.

    Uses relative offsets of labels from the block's top-left corner,
    sorted by (row_offset, col_offset, text), so two blocks with the
    same internal layout fingerprint identically regardless of their
    absolute position on the sheet.
    """
    if not label_cells:
        return hashlib.sha1(b"empty").hexdigest()[:16]

    parts: list[str] = []
    for c in label_cells:
        rel_row = c.row - top_row
        rel_col = c.col - top_col
        text = str(c.value or "")
        parts.append(f"{rel_row},{rel_col},{text}")
    parts.sort()
    joined = "|".join(parts).encode("utf-8")
    return hashlib.sha1(joined).hexdigest()[:16]


def _build_block(
    sheet: SheetSnapshot,
    bounds: _ComponentBounds,
    comp_id: int,
) -> Block | None:
    """
    Convert a component's bounding box into a Block.

    Returns None if the component fails the minimum-size filters
    (MIN_BLOCK_CELLS or MIN_LABEL_CELLS in config).
    """
    # Convert 0-indexed bounds → 1-indexed Excel coords
    top_row = bounds.top + 1
    top_col = bounds.left + 1
    bottom_row = bounds.bottom + 1
    bottom_col = bounds.right + 1

    # Span check — warn (don't split) if dilation over-merged
    if bounds.row_span > config.MAX_BLOCK_ROW_SPAN or bounds.col_span > config.MAX_BLOCK_COL_SPAN:
        logger.warning(
            "Sheet %r: block spanning rows %d-%d, cols %d-%d exceeds MAX_BLOCK_*_SPAN; "
            "consider lowering GAP_TOLERANCE.",
            sheet.name, top_row, bottom_row, top_col, bottom_col,
        )

    # Collect cells inside the bounding box (from ORIGINAL map, not dilated grid)
    contained: list[Cell] = []
    for cell in sheet.cells.values():
        if top_row <= cell.row <= bottom_row and top_col <= cell.col <= bottom_col:
            contained.append(cell)

    if len(contained) < config.MIN_BLOCK_CELLS:
        return None

    label_cells = [c for c in contained if c.dtype == CellType.LABEL]
    value_cells = [c for c in contained if c.dtype != CellType.LABEL]

    if len(label_cells) < config.MIN_LABEL_CELLS:
        return None

    primary_label = _pick_primary_label(label_cells)
    fingerprint = _compute_fingerprint(label_cells, top_row, top_col)

    block_id = f"{sheet.name}_R{top_row}C{top_col}_#{comp_id}"

    return Block(
        block_id=block_id,
        sheet_name=sheet.name,
        top_left=(top_row, top_col),
        bottom_right=(bottom_row, bottom_col),
        shape=(bottom_row - top_row + 1, bottom_col - top_col + 1),
        cell_count=len(contained),
        labels=[str(c.value) for c in label_cells if c.value is not None],
        primary_label=primary_label,
        label_cells=label_cells,
        value_cells=value_cells,
        fingerprint=fingerprint,
        neighbor_context=[],   # filled in after all blocks built
    )


def _compute_neighbor_context(blocks: list[Block]) -> None:
    """
    Populate the neighbor_context field on each block.

    A block's neighbors are other blocks whose bounding boxes lie within
    NEIGHBOR_RADIUS_ROWS / NEIGHBOR_RADIUS_COLS of this block's box.
    Mutates the blocks in place.
    """
    r_radius = config.NEIGHBOR_RADIUS_ROWS
    c_radius = config.NEIGHBOR_RADIUS_COLS

    for i, blk in enumerate(blocks):
        t1, l1 = blk.top_left
        b1, r1 = blk.bottom_right
        neighbors: list[tuple[int, int, str]] = []   # (row_distance, col_distance, label)

        for j, other in enumerate(blocks):
            if i == j or other.primary_label is None:
                continue
            t2, l2 = other.top_left
            b2, r2 = other.bottom_right

            # Gap between boxes (0 if they overlap/touch)
            row_gap = max(0, max(t1, t2) - min(b1, b2))
            col_gap = max(0, max(l1, l2) - min(r1, r2))

            if row_gap <= r_radius and col_gap <= c_radius:
                neighbors.append((row_gap + col_gap, t2, other.primary_label))

        # Sort by proximity, then by position, and keep labels
        neighbors.sort(key=lambda x: (x[0], x[1]))
        blk.neighbor_context = [lbl for _, _, lbl in neighbors]


# ── Public API ────────────────────────────────────────────────────────


def _components_should_merge(
    b1: _ComponentBounds,
    b2: _ComponentBounds,
    tolerance: int,
) -> bool:
    """
    Decide whether two tight components should be merged into one block.

    Merge criterion: the bounding boxes overlap along one axis AND are
    separated by at most `tolerance` empty cells along the perpendicular
    axis. Diagonal proximity alone is NOT sufficient — this avoids
    merging unrelated clusters that happen to be near each other on
    sparse sheets.
    """
    # Vertical adjacency: column ranges overlap, row gap within tolerance
    col_overlap = not (b1.right < b2.left or b2.right < b1.left)
    if col_overlap:
        row_gap = max(0, max(b1.top, b2.top) - min(b1.bottom, b2.bottom) - 1)
        if row_gap <= tolerance:
            return True

    # Horizontal adjacency: row ranges overlap, column gap within tolerance
    row_overlap = not (b1.bottom < b2.top or b2.bottom < b1.top)
    if row_overlap:
        col_gap = max(0, max(b1.left, b2.left) - min(b1.right, b2.right) - 1)
        if col_gap <= tolerance:
            return True

    return False


def _merge_close_components(
    bounds_map: dict[int, _ComponentBounds],
    tolerance: int,
) -> dict[int, _ComponentBounds]:
    """
    Merge tight components whose bounding boxes pass _components_should_merge.

    Uses union-find so that transitive merges are handled correctly: if
    A merges with B and B with C, all three end up in the same block
    even if A and C wouldn't merge directly. Returns a new dict mapping
    the union-find root id to the merged bounding box.
    """
    ids = sorted(bounds_map.keys())
    parent: dict[int, int] = {i: i for i in ids}

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]  # path compression
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    # Iterate pairs — with ~70 templates and typical sheets, component
    # counts stay well within O(n²) territory.
    for i, id1 in enumerate(ids):
        b1 = bounds_map[id1]
        for id2 in ids[i + 1:]:
            b2 = bounds_map[id2]
            if _components_should_merge(b1, b2, tolerance):
                union(id1, id2)

    # Collect components by root and compute their unioned bounding box
    groups: dict[int, list[_ComponentBounds]] = {}
    for i in ids:
        groups.setdefault(find(i), []).append(bounds_map[i])

    merged: dict[int, _ComponentBounds] = {}
    for root, boxes in groups.items():
        merged[root] = _ComponentBounds(
            top=min(b.top for b in boxes),
            left=min(b.left for b in boxes),
            bottom=max(b.bottom for b in boxes),
            right=max(b.right for b in boxes),
        )
    return merged


def detect_blocks_for_sheet(sheet: SheetSnapshot) -> list[Block]:
    """
    Detect blocks in a single sheet and return them.

    The sheet's `blocks` field is NOT mutated — caller decides whether
    to assign. Use detect_blocks_for_workbook() for in-place updates
    across the whole workbook.
    """
    if not sheet.cells or sheet.max_row == 0 or sheet.max_col == 0:
        return []

    grid = _build_occupancy_grid(sheet)
    if grid.size == 0:
        return []

    # Step 1: tight 8-connectivity labeling (no gap tolerance here)
    structure = np.ones((3, 3), dtype=np.uint8)
    labeled, num_components = nd_label(grid, structure=structure)
    bounds_map = _compute_component_bounds(labeled, num_components)

    # Step 2: post-hoc merge of components within GAP_TOLERANCE
    if config.GAP_TOLERANCE > 0 and len(bounds_map) > 1:
        bounds_map = _merge_close_components(bounds_map, config.GAP_TOLERANCE)

    # Step 3: build Block objects from final (merged) bounding boxes
    blocks: list[Block] = []
    for comp_id, bounds in bounds_map.items():
        block = _build_block(sheet, bounds, comp_id)
        if block is not None:
            blocks.append(block)

    # Sort in reading order for stable output
    blocks.sort(key=lambda b: (b.top_left[0], b.top_left[1]))

    _compute_neighbor_context(blocks)
    return blocks


def detect_blocks_for_workbook(snapshot: WorkbookSnapshot) -> int:
    """
    Detect blocks across every sheet in a workbook snapshot.

    Mutates the snapshot in place: each sheet's `blocks` list is
    populated, and `snapshot.total_blocks` is updated. Returns the
    total block count for convenience.

    Sheets listed in config.SKIP_BLOCK_DETECTION_SHEETS are bypassed —
    their `blocks` list is left empty. Cells, merges, and named ranges
    from those sheets are still available on the snapshot.
    """
    total = 0
    skip_set = config.SKIP_BLOCK_DETECTION_SHEETS
    for sheet_name, sheet in snapshot.sheets.items():
        if sheet_name in skip_set:
            sheet.blocks = []
            logger.info(
                "Skipped block detection for sheet %r (in SKIP_BLOCK_DETECTION_SHEETS)",
                sheet_name,
            )
            continue
        blocks = detect_blocks_for_sheet(sheet)
        sheet.blocks = blocks
        total += len(blocks)
        logger.info(
            "Detected %d blocks in sheet %r (%d cells)",
            len(blocks), sheet_name, len(sheet.cells),
        )
    snapshot.total_blocks = total
    return total


# ── Debug visualization ───────────────────────────────────────────────


def _block_overlay_for_sheet(sheet: SheetSnapshot, max_rows: int = 60, max_cols: int = 30) -> str:
    """
    Render the sheet as an ASCII grid with block boundaries overlaid.

    Each block gets a letter (A, B, C, ... AA, AB, ...). Non-empty cells
    inside a block show that letter; cells outside any block show '.'
    Empty cells are blank.

    Large sheets are truncated to max_rows × max_cols with an ellipsis
    marker; the accompanying block legend is always complete.
    """
    if sheet.max_row == 0 or sheet.max_col == 0:
        return f"[Sheet {sheet.name!r} is empty]\n"

    # Build lookup: (row, col) → block letter
    def _letter(i: int) -> str:
        """0 → 'A', 25 → 'Z', 26 → 'AA', ..."""
        result = ""
        i += 1
        while i > 0:
            i, rem = divmod(i - 1, 26)
            result = chr(ord("A") + rem) + result
        return result

    cell_to_letter: dict[tuple[int, int], str] = {}
    legend_lines: list[str] = []

    for idx, blk in enumerate(sheet.blocks):
        letter = _letter(idx)
        for cell in blk.label_cells + blk.value_cells:
            cell_to_letter[(cell.row, cell.col)] = letter
        tl = f"{_col_letter(blk.top_left[1])}{blk.top_left[0]}"
        br = f"{_col_letter(blk.bottom_right[1])}{blk.bottom_right[0]}"
        primary = blk.primary_label or "(no label)"
        legend_lines.append(
            f"  {letter}: {primary!r:40s}  {tl}:{br}  "
            f"shape={blk.shape}  cells={blk.cell_count}  fp={blk.fingerprint}"
        )

    # Render grid
    shown_rows = min(sheet.max_row, max_rows)
    shown_cols = min(sheet.max_col, max_cols)

    lines: list[str] = []
    lines.append(f"Sheet: {sheet.name}  ({sheet.max_row}x{sheet.max_col}, "
                 f"{len(sheet.cells)} cells, {len(sheet.blocks)} blocks)")
    lines.append("")

    # Column header: show column letters every 5 cols
    header = "     " + "".join(
        (_col_letter(c + 1)[-1] if (c + 1) % 5 == 0 or c == 0 else " ")
        for c in range(shown_cols)
    )
    lines.append(header)
    lines.append("    +" + "-" * shown_cols + "+")

    for r in range(shown_rows):
        row_str = []
        for c in range(shown_cols):
            addr = f"{_col_letter(c + 1)}{r + 1}"
            if addr in sheet.cells:
                letter = cell_to_letter.get((r + 1, c + 1), "?")
                # For multi-char letters, just use the last char in-grid
                row_str.append(letter[-1])
            else:
                row_str.append(" ")
        lines.append(f"{r + 1:4d}|" + "".join(row_str) + "|")

    lines.append("    +" + "-" * shown_cols + "+")

    if sheet.max_row > max_rows or sheet.max_col > max_cols:
        lines.append(
            f"    (truncated — full sheet is {sheet.max_row}x{sheet.max_col})"
        )

    lines.append("")
    lines.append("Blocks:")
    if legend_lines:
        lines.extend(legend_lines)
    else:
        lines.append("  (none detected)")
    lines.append("")

    return "\n".join(lines)


def _col_letter(col: int) -> str:
    """1 → 'A', 26 → 'Z', 27 → 'AA', ..."""
    result = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(ord("A") + rem) + result
    return result


def write_debug_report(
    snapshot: WorkbookSnapshot,
    output_path: Path | None = None,
    max_rows_per_sheet: int = 60,
    max_cols_per_sheet: int = 30,
) -> Path:
    """
    Write a human-readable block-detection debug report.

    Produces one text file covering every sheet in the snapshot, with
    an ASCII overlay of detected blocks and a legend showing each
    block's primary label, position, shape, and fingerprint.

    Use this after running detect_blocks_for_workbook() to tune
    GAP_TOLERANCE and the other block detection parameters on real
    data. If the overlays merge things that shouldn't be merged, lower
    GAP_TOLERANCE. If they split things that belong together, raise it.

    Args:
        snapshot: Workbook snapshot with blocks already populated.
        output_path: Destination file. Defaults to
            config.REPORT_DIR / "<file_stem>_blocks.txt".
        max_rows_per_sheet: Truncate overlay grids past this row count.
        max_cols_per_sheet: Truncate overlay grids past this column count.

    Returns:
        Path to the written report file.
    """
    if output_path is None:
        stem = Path(snapshot.file_name).stem
        output_path = config.REPORT_DIR / f"{stem}_blocks.txt"
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sections: list[str] = []
    sections.append("=" * 78)
    sections.append(f"BLOCK DETECTION DEBUG REPORT")
    sections.append("=" * 78)
    sections.append(f"File:         {snapshot.file_name}")
    sections.append(f"Month:        {snapshot.month_label or '(unlabeled)'}")
    sections.append(f"Sheets:       {len(snapshot.sheet_names)}")
    sections.append(f"Total cells:  {snapshot.total_cells}")
    sections.append(f"Total blocks: {snapshot.total_blocks}")
    sections.append(f"GAP_TOLERANCE: {config.GAP_TOLERANCE}   "
                    f"MIN_BLOCK_CELLS: {config.MIN_BLOCK_CELLS}   "
                    f"MIN_LABEL_CELLS: {config.MIN_LABEL_CELLS}")
    sections.append("=" * 78)
    sections.append("")

    for sheet_name in snapshot.sheet_names:
        sheet = snapshot.sheets[sheet_name]
        sections.append("-" * 78)
        sections.append(_block_overlay_for_sheet(
            sheet, max_rows=max_rows_per_sheet, max_cols=max_cols_per_sheet
        ))

    output_path.write_text("\n".join(sections), encoding="utf-8")
    logger.info("Wrote block debug report: %s", output_path)
    return output_path
